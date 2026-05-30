# -*- coding: utf-8 -*-
"""
Parallel runner for GitHub Actions.
Uses subprocess to launch workers sequentially-staggered.
"""

import os
import sys
import json
import math
import time
import datetime
import subprocess
import csv
import openpyxl

BASE_URL    = "https://www.thegreyhoundrecorder.com.au"
NUM_WORKERS = int(os.environ.get("NUM_WORKERS", "3"))
INPUT_FILE  = os.environ.get("INPUT_FILE", "input.json")
OUTPUT_DIR  = os.path.join(os.path.dirname(__file__), "..", "output")

os.makedirs(OUTPUT_DIR, exist_ok=True)


def excel_serial_to_date(serial):
    base = datetime.date(1899, 12, 30)
    return base + datetime.timedelta(days=int(serial))


def read_excel_input(path):
    rows = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        date_val = row[0]
        dog_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if not dog_name:
            continue
        if isinstance(date_val, datetime.datetime):
            d = date_val.date()
        elif isinstance(date_val, datetime.date):
            d = date_val
        elif isinstance(date_val, (int, float)):
            d = excel_serial_to_date(date_val)
        else:
            continue
        rows.append({"date": d.strftime("%Y-%m-%d"), "dog": dog_name})
    wb.close()
    return rows


def split_chunks(items, n):
    size = math.ceil(len(items) / n)
    return [items[i:i+size] for i in range(0, len(items), size)]


def merge_to_excel(csv_files, output_path, original_order):
    from openpyxl.styles import Font, Border, Side
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    order_map = {}
    for idx, item in enumerate(original_order):
        key = (item["date"], item["dog"].lower())
        order_map[key] = idx

    header = None
    all_rows = []
    for path in csv_files:
        if not os.path.exists(path):
            print(f"Missing: {path}", flush=True)
            continue
        with open(path, newline="", encoding="utf-8-sig") as fin:
            reader = csv.reader(fin)
            rows = list(reader)
            if not rows:
                continue
            if header is None:
                header = rows[0]
            for row in rows[1:]:
                if row:
                    all_rows.append(row)

    def sort_key(row):
        date_val = row[0] if len(row) > 0 else ""
        dog_val  = row[1].lower() if len(row) > 1 else ""
        return order_map.get((date_val, dog_val), 999999)

    all_rows.sort(key=sort_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    row_num = 1

    if header:
        for c, val in enumerate(header, 1):
            cell = ws.cell(row=row_num, column=c, value=val)
            cell.font = Font(bold=True, size=10)
            cell.border = border
        row_num += 1

    for row in all_rows:
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=c, value=val)
            cell.font = Font(size=10)
            cell.border = border
        row_num += 1

    wb.save(output_path)
    print(f"Merged {len(csv_files)} files, {len(all_rows)} rows -> {output_path}", flush=True)


def main():
    # Load input
    input_path = os.path.join(os.path.dirname(__file__), "..", INPUT_FILE)
    if input_path.endswith(".json"):
        with open(input_path) as f:
            items = json.load(f)
    else:
        items = read_excel_input(input_path)

    if not items:
        print("No input rows found.", flush=True)
        return

    print(f"Loaded {len(items)} rows. Running {NUM_WORKERS} parallel workers.", flush=True)

    chunks = split_chunks(items, NUM_WORKERS)
    chunk_files  = []
    output_csvs  = []
    processes    = []

    worker_script = os.path.join(os.path.dirname(__file__), "worker.py")

    # Write chunks and launch workers with stagger
    for i, chunk in enumerate(chunks):
        chunk_file = os.path.join(OUTPUT_DIR, f"chunk_{i}.json")
        output_csv = os.path.join(OUTPUT_DIR, f"results_worker_{i}.csv")
        with open(chunk_file, "w") as f:
            json.dump(chunk, f)
        chunk_files.append(chunk_file)
        output_csvs.append(output_csv)

        print(f"\nStarting worker {i} ({len(chunk)} rows)...", flush=True)
        env = os.environ.copy()
        p = subprocess.Popen(
            [sys.executable, worker_script, str(i), chunk_file, output_csv],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            env=env
        )
        processes.append((i, p))

        # Stagger: wait for Chrome to fully initialise before next worker
        # This prevents uc_driver file conflicts
        if i < len(chunks) - 1:
            print(f"  Waiting 15s before starting worker {i+1}...", flush=True)
            time.sleep(15)

    # Stream output from all workers
    print("\n--- Worker output ---", flush=True)
    import threading

    def stream(worker_id, proc):
        for line in proc.stdout:
            print(f"[W{worker_id}] {line}", end="", flush=True)
        proc.wait()
        print(f"[W{worker_id}] Exit code: {proc.returncode}", flush=True)

    threads = []
    for worker_id, p in processes:
        t = threading.Thread(target=stream, args=(worker_id, p), daemon=True)
        t.start()
        threads.append(t)

    for t in threads:
        t.join()

    print("\n--- All workers done ---", flush=True)

    # Merge results
    merged_path = os.path.join(OUTPUT_DIR, "results_merged.xlsx")
    merge_to_excel(output_csvs, merged_path, items)

    # Collect failed
    all_failed = []
    for i in range(len(chunks)):
        failed_path = os.path.join(OUTPUT_DIR, f"results_worker_{i}_failed.json")
        if os.path.exists(failed_path):
            with open(failed_path) as f:
                all_failed.extend(json.load(f))

    if all_failed:
        with open(os.path.join(OUTPUT_DIR, "failed.json"), "w") as f:
            json.dump(all_failed, f, indent=2)
        print(f"\n{len(all_failed)} rows failed — saved to failed.json", flush=True)

    # Cleanup
    for cf in chunk_files:
        try: os.remove(cf)
        except: pass

    print(f"\nDone. Results at: {merged_path}", flush=True)


if __name__ == "__main__":
    main()
